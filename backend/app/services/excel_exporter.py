"""
Excel export service for formatting and exporting event data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Optional
from datetime import datetime
from io import BytesIO
from loguru import logger

from app.models import EventData, EventType


class ExcelExporter:
    """
    Service for exporting event data to formatted Excel files.
    Creates professional-looking Excel workbooks with proper styling.
    """
    
    # Color scheme
    HEADER_COLOR = "366092"  # Dark blue
    ALT_ROW_COLOR = "F2F2F2"  # Light gray
    LINK_COLOR = "0563C1"    # Blue for hyperlinks
    
    def __init__(self):
        """Initialize the Excel exporter."""
        logger.info("ExcelExporter initialized")
    
    def _create_header_style(self) -> dict:
        """
        Create header cell styling.
        
        Returns:
            Dictionary of style attributes
        """
        return {
            'font': Font(bold=True, color="FFFFFF", size=11),
            'fill': PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid"),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def _create_cell_style(self, is_alt_row: bool = False) -> dict:
        """
        Create data cell styling.
        
        Args:
            is_alt_row: Whether this is an alternating row (for zebra striping)
        
        Returns:
            Dictionary of style attributes
        """
        style = {
            'alignment': Alignment(vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
        }
        
        if is_alt_row:
            style['fill'] = PatternFill(start_color=self.ALT_ROW_COLOR, end_color=self.ALT_ROW_COLOR, fill_type="solid")
        
        return style
    
    def _apply_style(self, cell, style_dict: dict):
        """
        Apply style dictionary to a cell.
        
        Args:
            cell: Excel cell object
            style_dict: Dictionary of style attributes
        """
        for attr, value in style_dict.items():
            setattr(cell, attr, value)
    
    def _auto_adjust_column_widths(self, worksheet, min_width: int = 10, max_width: int = 50):
        """
        Auto-adjust column widths based on content.
        
        Args:
            worksheet: Excel worksheet
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_list(self, items: List[str]) -> str:
        """
        Format a list of strings for Excel display.
        
        Args:
            items: List of strings
        
        Returns:
            Formatted string
        """
        if not items:
            return ""
        return ", ".join(items)
    
    def _format_date(self, date: datetime) -> str:
        """
        Format datetime for Excel display.
        
        Args:
            date: Datetime object
        
        Returns:
            Formatted date string
        """
        if not date:
            return ""
        return date.strftime("%Y-%m-%d %H:%M")
    
    def create_events_workbook(
        self,
        events: List[EventData],
        include_metadata: bool = True
    ) -> Workbook:
        """
        Create an Excel workbook with event data.
        
        Args:
            events: List of EventData objects
            include_metadata: Whether to include a metadata sheet
        
        Returns:
            Workbook object
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create main events sheet
        self._create_events_sheet(wb, events)
        
        # Create summary sheet
        if include_metadata:
            self._create_summary_sheet(wb, events)
        
        logger.info(f"Created Excel workbook with {len(events)} events")
        return wb
    
    def _create_events_sheet(self, workbook: Workbook, events: List[EventData]):
        """
        Create the main events data sheet.
        
        Args:
            workbook: Workbook object
            events: List of EventData objects
        """
        ws = workbook.create_sheet("Events", 0)
        
        # Define headers
        headers = [
            "Event Type",
            "Title",
            "Summary",
            "Location",
            "Date/Time",
            "Participants",
            "Organizations",
            "Confidence",
            "Source URL",
            "Full Context"
        ]
        
        # Write headers
        header_style = self._create_header_style()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, header_style)
        
        # Write data rows
        for row_idx, event in enumerate(events, 2):
            is_alt_row = (row_idx % 2) == 0
            cell_style = self._create_cell_style(is_alt_row)
            
            # Event Type
            cell = ws.cell(row=row_idx, column=1, value=event.event_type.value.upper())
            self._apply_style(cell, cell_style)
            
            # Title
            cell = ws.cell(row=row_idx, column=2, value=event.title)
            self._apply_style(cell, cell_style)
            cell.font = Font(bold=True)
            
            # Summary
            cell = ws.cell(row=row_idx, column=3, value=event.summary)
            self._apply_style(cell, cell_style)
            
            # Location
            location_str = str(event.location) if event.location else ""
            cell = ws.cell(row=row_idx, column=4, value=location_str)
            self._apply_style(cell, cell_style)
            
            # Date/Time
            date_str = self._format_date(event.event_date)
            cell = ws.cell(row=row_idx, column=5, value=date_str)
            self._apply_style(cell, cell_style)
            
            # Participants
            participants_str = self._format_list(event.participants)
            cell = ws.cell(row=row_idx, column=6, value=participants_str)
            self._apply_style(cell, cell_style)
            
            # Organizations
            orgs_str = self._format_list(event.organizations)
            cell = ws.cell(row=row_idx, column=7, value=orgs_str)
            self._apply_style(cell, cell_style)
            
            # Confidence
            confidence_str = f"{event.confidence:.0%}"
            cell = ws.cell(row=row_idx, column=8, value=confidence_str)
            self._apply_style(cell, cell_style)
            
            # Source URL (with hyperlink)
            if event.source_url:
                cell = ws.cell(row=row_idx, column=9, value=event.source_url)
                cell.hyperlink = event.source_url
                cell.font = Font(color=self.LINK_COLOR, underline="single")
                self._apply_style(cell, cell_style)
            else:
                cell = ws.cell(row=row_idx, column=9, value="")
                self._apply_style(cell, cell_style)
            
            # Full Context (cleaned article content)
            full_context = event.full_content if event.full_content else ""
            # Clean up the text - remove excessive whitespace
            full_context = " ".join(full_context.split())
            cell = ws.cell(row=row_idx, column=10, value=full_context)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            self._apply_style(cell, cell_style)
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
        
        # Freeze top row
        ws.freeze_panes = "A2"
        
        logger.info(f"Created Events sheet with {len(events)} rows")
    
    def _create_summary_sheet(self, workbook: Workbook, events: List[EventData]):
        """
        Create a summary/metadata sheet.
        
        Args:
            workbook: Workbook object
            events: List of EventData objects
        """
        ws = workbook.create_sheet("Summary", 1)
        
        # Title
        ws['A1'] = "Event Export Summary"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Export info
        row = 3
        ws[f'A{row}'] = "Export Date:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Total Events:"
        ws[f'B{row}'] = len(events)
        ws[f'B{row}'].font = Font(bold=True)
        
        # Event type breakdown
        row += 2
        ws[f'A{row}'] = "Event Type Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        # Count events by type
        type_counts = {}
        for event in events:
            event_type = event.event_type.value
            type_counts[event_type] = type_counts.get(event_type, 0) + 1
        
        row += 1
        ws[f'A{row}'] = "Event Type"
        ws[f'B{row}'] = "Count"
        header_style = self._create_header_style()
        self._apply_style(ws[f'A{row}'], header_style)
        self._apply_style(ws[f'B{row}'], header_style)
        
        for event_type, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            row += 1
            ws[f'A{row}'] = event_type.upper()
            ws[f'B{row}'] = count
        
        # Location breakdown
        row += 2
        ws[f'A{row}'] = "Top Locations"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        # Count events by location
        location_counts = {}
        for event in events:
            if event.location and event.location.country:
                country = event.location.country
                location_counts[country] = location_counts.get(country, 0) + 1
        
        row += 1
        ws[f'A{row}'] = "Country"
        ws[f'B{row}'] = "Count"
        self._apply_style(ws[f'A{row}'], header_style)
        self._apply_style(ws[f'B{row}'], header_style)
        
        for country, count in sorted(location_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
            row += 1
            ws[f'A{row}'] = country
            ws[f'B{row}'] = count
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
        
        logger.info("Created Summary sheet")
    
    def export_to_bytes(
        self,
        events: List[EventData],
        include_metadata: bool = True
    ) -> BytesIO:
        """
        Export events to Excel file in memory (BytesIO).
        
        Args:
            events: List of EventData objects
            include_metadata: Whether to include metadata sheet
        
        Returns:
            BytesIO object containing Excel file
        """
        if not events:
            logger.warning("Attempted to export empty event list")
            raise ValueError("Cannot export empty event list")
        
        # Create workbook
        wb = self.create_events_workbook(events, include_metadata)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Exported {len(events)} events to Excel (BytesIO)")
        return output
    
    def export_to_file(
        self,
        events: List[EventData],
        filepath: str,
        include_metadata: bool = True
    ):
        """
        Export events to Excel file on disk.
        
        Args:
            events: List of EventData objects
            filepath: Path to save Excel file
            include_metadata: Whether to include metadata sheet
        """
        if not events:
            logger.warning("Attempted to export empty event list")
            raise ValueError("Cannot export empty event list")
        
        # Create workbook
        wb = self.create_events_workbook(events, include_metadata)
        
        # Save to file
        wb.save(filepath)
        
        logger.info(f"Exported {len(events)} events to {filepath}")
    
    def get_default_filename(self) -> str:
        """
        Generate a default filename for Excel export.
        
        Returns:
            Filename string with timestamp
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"events_export_{timestamp}.xlsx"


# Global exporter instance
excel_exporter = ExcelExporter()
